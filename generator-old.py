import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import numpy as np
import re
import os
from datetime import datetime

# --- Configuration ---
EXCLUDED_TAGS = {
    "Female Protagonist", "Male Protagonist", "Primarily Female Cast", 
    "Primarily Male Cast", "School", "Heterosexual", "Primarily Teen Cast",
    "Ensemble Cast"
}

def extract_year(vintage_str):
    if not vintage_str: return None
    years = re.findall(r'\d{4}', str(vintage_str))
    if not years: return None
    year_val = float(years[0])
    season_map = {"winter": 0.00, "spring": 0.25, "summer": 0.50, "fall": 0.75}
    v_lower = str(vintage_str).lower()
    decimal = 0.0
    for season, val in season_map.items():
        if season in v_lower:
            decimal = val
            break
    return year_val + decimal

# --- UI COMPONENTS ---

class SubSelectionDialog(tk.Toplevel):
    def __init__(self, parent, missing_roster):
        super().__init__(parent)
        self.title("Substitute Resolution")
        self.result = None
        tk.Label(self, text="Multiple roster members are missing.\nWhich player is being replaced by the substitute?", 
                 font=("Arial", 10), padx=20, pady=10).pack()
        self.listbox = tk.Listbox(self, height=len(missing_roster))
        self.listbox.pack(padx=20, pady=5, fill=tk.X)
        for m in missing_roster: self.listbox.insert(tk.END, m)
        ttk.Button(self, text="Confirm", command=self.on_confirm).pack(pady=10)
        self.grab_set(); self.wait_window()
        
    def on_confirm(self):
        sel = self.listbox.curselection()
        if sel: self.result = self.listbox.get(sel[0]); self.destroy()

class ManualMatchDialog(tk.Toplevel):
    def __init__(self, parent, unknown_name, available_pool):
        super().__init__(parent)
        self.title("Manual Match Required")
        self.result = None
        ttk.Label(self, text=f"Could not find match for: '{unknown_name}'", font=("Arial", 10, "bold")).pack(pady=10)
        self.listbox = tk.Listbox(self, height=15); self.listbox.pack(padx=10, fill=tk.BOTH)
        for name in sorted(available_pool): self.listbox.insert(tk.END, name)
        ttk.Button(self, text="Match Selected", command=self.on_match).pack(pady=10)
        self.grab_set(); self.wait_window()

    def on_match(self):
        sel = self.listbox.curselection()
        if sel: self.result = self.listbox.get(sel[0]); self.destroy()

# --- CORE LOGIC ---

def process_files():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_dir = os.path.join(script_dir, "jsons")
    
    # 1. Robust JSON Discovery with Retry Loop
    json_paths = []
    while True:
        if os.path.exists(json_dir) and os.path.isdir(json_dir):
            json_paths = [os.path.join(json_dir, f) for f in os.listdir(json_dir) if f.endswith(".json")]
        
        if json_paths:
            break
        else:
            retry = messagebox.askyesno("Missing Files", "There is no jsons folder detected or there are no JSON files in the folder. Lock in and press yes to re-run the script")
            if not retry:
                return

    all_known_players = set()
    for path in json_paths:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f); songs = data.get("songs", [])
                for s in songs:
                    for p in s.get("correctGuessPlayers", []): all_known_players.add(p)
                    for ls in s.get("listStates", []): all_known_players.add(ls["name"])
        except: continue

    raw_assignments = {}
    team_rosters = defaultdict(set)
    t1_lookup = {}
    use_teams = False

    # 2. Team Assignment with Empty/Missing Check
    codes_path = os.path.join(script_dir, "dependencies", "codes.txt")
    codes_valid = False
    if os.path.exists(codes_path):
        with open(codes_path, "r", encoding="utf-8") as f:
            if f.read().strip():
                codes_valid = True

    if not codes_valid:
        if not messagebox.askyesno("Codes Missing", "codes.txt is missing or empty, skip team assignment phase?"):
            return
    else:
        with open(codes_path, "r", encoding="utf-8") as f:
            content = f.read()
        all_teams_data = []
        for line in content.strip().split('\n'):
            matches = re.findall(r'([^\s(]+)\s*\([\d.]+\)', line)
            if matches: all_teams_data.append(matches[:4])

        if all_teams_data:
            use_teams = True
            available = list(all_known_players)
            for t_idx, members in enumerate(all_teams_data, 1):
                for i, p_in in enumerate(members):
                    tier = f"T{i+1}"
                    match = next((n for n in available if n.lower() == p_in.lower()), None)
                    if not match:
                        d = ManualMatchDialog(None, p_in, available)
                        match = d.result
                    if match:
                        raw_assignments[match] = (t_idx, tier)
                        team_rosters[t_idx].add(match)
                        if match in available: available.remove(match)
                        if tier == "T1": t1_lookup[t_idx] = match

    # Statistical Counters
    correct_counts, song_participation = defaultdict(int), defaultdict(int)
    erigs_counts, player_reverse_erigs = defaultdict(int), defaultdict(int)
    player_two_eighths, player_points, player_blocks = defaultdict(int), defaultdict(int), defaultdict(int)
    player_type_correct, player_type_seen = defaultdict(lambda: defaultdict(int)), defaultdict(lambda: defaultdict(int))
    player_rigs, player_rigs_hit = defaultdict(int), defaultdict(int)
    all_song_vintages, all_song_difficulties = [], []
    total_correct_answers_sum, total_erigs = 0, 0
    genre_counter, tag_counter = Counter(), Counter()
    player_list_vintages, player_list_correct_counts = defaultdict(list), defaultdict(list) 
    player_missed_erigs, watched_only_valid = defaultdict(int), False
    team_correct_per_song = defaultdict(list)
    team_onlist_synergy, team_offlist_synergy, team_shared_rig_pct = defaultdict(list), defaultdict(list), defaultdict(list)

    for path in json_paths:
        with open(path, encoding="utf-8") as f: data = json.load(f); songs = data.get("songs", [])
        if not songs: continue
        
        raw_file_players = set()
        for song in songs:
            for p in song.get("correctGuessPlayers", []): raw_file_players.add(p)
            for ls in song.get("listStates", []): raw_file_players.add(ls["name"])
        
        final_file_members = set(raw_file_players)
        if use_teams:
            teams_in_file = set(raw_assignments[p][0] for p in raw_file_players if p in raw_assignments)
            for t_id in teams_in_file:
                roster = team_rosters[t_id]
                missing = [p for p in roster if p not in raw_file_players]
                if len([p for p in roster if p in raw_file_players]) == 3 and missing:
                    if len(missing) == 1: final_file_members.add(missing[0])
                    else:
                        d = SubSelectionDialog(None, missing)
                        if d.result: final_file_members.add(d.result)

        apply_rev = (len(final_file_members) % 2 == 0)
        max_songs = max(s.get("songNumber", 0) for s in songs)
        type_totals_this_file = defaultdict(int)

        for song in songs:
            si = song.get("songInfo", {}); st = si.get("type")
            if st in [1, 2, 3]: type_totals_this_file[st] += 1
            if isinstance(si.get("animeGenre"), list): genre_counter.update(si.get("animeGenre"))
            if isinstance(si.get("animeTags"), list):
                tag_counter.update([t for t in si.get("animeTags") if t not in EXCLUDED_TAGS])

            correct = set(song.get("correctGuessPlayers", []))
            ls = song.get("listStates", []); total_correct_answers_sum += len(correct)
            year, diff = extract_year(si.get("vintage")), si.get("animeDifficulty")
            if isinstance(diff, (int, float)): all_song_difficulties.append(diff)
            if year is not None: all_song_vintages.append(year)
            
            song_riggers = {p["name"] for p in ls}
            
            if use_teams:
                teams_in_this_file = list(set(raw_assignments[p][0] for p in raw_file_players if p in raw_assignments))
                if len(teams_in_this_file) == 2:
                    tA, tB = teams_in_this_file[0], teams_in_this_file[1]
                    for cur_t, opp_t in [(tA, tB), (tB, tA)]:
                        cur_correct = correct.intersection(team_rosters[cur_t])
                        opp_correct = correct.intersection(team_rosters[opp_t])
                        if not opp_correct:
                            for p in cur_correct: player_points[p] += 1
                        if len(cur_correct) == 1 and len(opp_correct) > 0:
                            player_blocks[list(cur_correct)[0]] += 1

                for t_id in teams_in_this_file:
                    roster = team_rosters[t_id]
                    correct_on_team = correct.intersection(roster)
                    team_correct_per_song[t_id].append(len(correct_on_team) / 4.0)
                    team_riggers = song_riggers.intersection(roster)
                    if team_riggers:
                        team_onlist_synergy[t_id].append(len(correct_on_team) / 4.0)
                        team_shared_rig_pct[t_id].append((len(team_riggers) - 1) / 3.0)
                    else: team_offlist_synergy[t_id].append(len(correct_on_team) / 4.0)

            if len(correct) == 2:
                for p in correct: player_two_eighths[p] += 1
            elif len(correct) == 1: 
                total_erigs += 1; erigs_counts[list(correct)[0]] += 1
            if apply_rev and len(final_file_members - correct) == 1:
                player_reverse_erigs[list(final_file_members - correct)[0]] += 1

            for name in final_file_members:
                if name in correct:
                    correct_counts[name] += 1
                    if st in [1, 2, 3]: player_type_correct[name][st] += 1
            if ls:
                watched_only_valid = True
                for p in ls:
                    n = p["name"]; player_rigs[n] += 1
                    if n in correct: player_rigs_hit[n] += 1
                    if year is not None: player_list_vintages[n].append(year)
                    player_list_correct_counts[n].append(len(correct))
                    if len(correct) == 0: player_missed_erigs[n] += 1

        for name in final_file_members:
            song_participation[name] += max_songs
            for t in [1, 2, 3]: player_type_seen[name][t] += type_totals_this_file[t]

    # --- DATAFRAME GENERATION ---
    p_rows = []
    for name in song_participation:
        total, correct = song_participation[name], correct_counts[name]
        t_id, tier = raw_assignments.get(name, ("Unassigned", "N/A"))
        t_name = t1_lookup.get(t_id, "Unknown")
        p_rows.append({
            "Team": t_name, "Tier": tier, "Player": name, 
            "Guess Rate": correct/total if total else 0, "Erigs 🔫": erigs_counts[name],
            "Points": player_points[name], "Blocks": player_blocks[name],
            "2/8s": player_two_eighths[name], "Rev. Erigs": player_reverse_erigs[name],
            "Song Count": total,
            "OP GR": player_type_correct[name][1]/player_type_seen[name][1] if player_type_seen[name][1] else np.nan,
            "ED GR": player_type_correct[name][2]/player_type_seen[name][2] if player_type_seen[name][2] else np.nan,
            "IN GR": player_type_correct[name][3]/player_type_seen[name][3] if player_type_seen[name][3] else np.nan,
            "Rigs": player_rigs[name], "Rigs Missed": player_rigs[name]-player_rigs_hit[name],
            "Onlist GR": player_rigs_hit[name]/player_rigs[name] if player_rigs[name] else np.nan,
            "Offlist GR": (correct-player_rigs_hit[name])/(total-player_rigs[name]) if (total-player_rigs[name]) else np.nan
        })
    df_ps = pd.DataFrame(p_rows).sort_values("Guess Rate", ascending=False)

    df_tour = pd.DataFrame([
        ["Average Vintage", round(np.mean(all_song_vintages), 2) if all_song_vintages else "N/A"],
        ["Average Difficulty", round(np.mean(all_song_difficulties), 2) if all_song_difficulties else "N/A"],
        ["Average GR", f"{(total_correct_answers_sum / sum(song_participation.values())):.2%}" if sum(song_participation.values()) > 0 else "0.00%"],
        ["Total Erigs", total_erigs],
        ["Total Rev. Erigs", sum(player_reverse_erigs.values())],
        ["Most Popular Genre", f"{genre_counter.most_common(1)[0][0]} ({genre_counter.most_common(1)[0][1]})" if genre_counter else "N/A"],
        ["Most Popular Tag", f"{tag_counter.most_common(1)[0][0]} ({tag_counter.most_common(1)[0][1]})" if tag_counter else "N/A"],
    ], columns=["TOUR STATS", ""])

    # --- FIX: Only process team stats if teams were assigned ---
    team_stat_rows, team_meta = [], []
    if use_teams:
        for t_id in sorted(team_correct_per_song.keys()):
            t_name = t1_lookup.get(t_id, f"Team {t_id}"); roster = team_rosters[t_id]
            t_v = [v for p in roster for v in player_list_vintages[p]]
            t_d = [v for p in roster for v in player_list_correct_counts[p]]
            team_stat_rows.append({"TEAM STATS": t_name, "Avg. Correct": np.mean(team_correct_per_song[t_id]), "Onlist Synergy": np.mean(team_onlist_synergy[t_id]) if team_onlist_synergy[t_id] else 0, "Offlist Synergy": np.mean(team_offlist_synergy[t_id]) if team_offlist_synergy[t_id] else 0, "Shared Rigs": np.mean(team_shared_rig_pct[t_id]) if team_shared_rig_pct[t_id] else 0})
            team_meta.append({"name": t_name, "erigs": sum(erigs_counts[p] for p in roster), "vintage": np.mean(t_v) if t_v else 0, "diff": np.mean(t_d) if t_d else 0})
    
    df_team_stats = pd.DataFrame(team_stat_rows)
    if not df_team_stats.empty:
        df_team_stats = df_team_stats.sort_values("Avg. Correct", ascending=False)

    tier_hero_rows = []
    if use_teams:
        for tier in ["T1", "T2", "T3", "T4"]:
            tp = [p for p, attr in raw_assignments.items() if attr[1] == tier]
            if tp:
                bp = max(tp, key=lambda x: player_points[x]); bb = max(tp, key=lambda x: player_blocks[x])
                tier_hero_rows.append([tier, f"{bp} ({player_points[bp]})", f"{bb} ({player_blocks[bb]})"])
    df_tier_heroes = pd.DataFrame(tier_hero_rows, columns=["Tier", "Top Attacker", "Top Blocker"])

    # --- EXPORT ---
    timestamp = datetime.now().strftime("%m%d%H%M")
    out_name = f"export_{timestamp}.xlsx"
    out_path = os.path.join(script_dir, out_name)
    
    df_display = df_ps.copy()
    pct_cols_p = ["Guess Rate", "OP GR", "ED GR", "IN GR", "Onlist GR", "Offlist GR"]
    for c in pct_cols_p: df_display[c] = df_display[c].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "N/A")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_display.to_excel(writer, sheet_name="Player Stats", index=False)
        df_tour.to_excel(writer, sheet_name="Extra Stats", index=False)
        
        # Only write team sections if they exist
        if use_teams and not df_team_stats.empty:
            df_team_display = df_team_stats.copy()
            pct_cols_t = ["Avg. Correct", "Onlist Synergy", "Offlist Synergy", "Shared Rigs"]
            for c in pct_cols_t: df_team_display[c] = df_team_display[c].apply(lambda x: f"{x:.2%}")
            df_team_display.to_excel(writer, sheet_name="Extra Stats", index=False, startcol=3)
            
            erig_r = len(df_team_stats) + 2
            if team_meta:
                m_e = sorted(team_meta, key=lambda x: x['erigs'], reverse=True)[0]
                m_z = sorted(team_meta, key=lambda x: x['vintage'], reverse=True)[0]
                m_b = sorted(team_meta, key=lambda x: x['vintage'])[0]
                m_ea = sorted(team_meta, key=lambda x: x['diff'], reverse=True)[0]
                m_h = sorted(team_meta, key=lambda x: x['diff'])[0]
                df_team_super = pd.DataFrame([
                    ["Team with the most erigs", f"{m_e['name']} ({m_e['erigs']})"],
                    ["Most zoomer team", f"{m_z['name']} ({round(m_z['vintage'], 2)})"],
                    ["Most boomer team", f"{m_b['name']} ({round(m_b['vintage'], 2)})"],
                    ["Team with the easiest lists", f"{m_ea['name']} ({round(m_ea['diff'], 2)})"],
                    ["Team with the hardest lists", f"{m_h['name']} ({round(m_h['diff'], 2)})"],
                ])
                df_team_super.iloc[0:1].to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=erig_r, startcol=3)
                df_team_super.iloc[1:].to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=erig_r + 1, startcol=3)
            
            if not df_tier_heroes.empty:
                df_tier_heroes.to_excel(writer, sheet_name="Extra Stats", index=False, startrow=erig_r + 6, startcol=3)

        base_r = len(df_tour) + 2
        pd.DataFrame([["WATCHED STATS"]]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r)
        
        if watched_only_valid:
            plist = list(song_participation.keys())
            e = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_correct_counts[n])) for n in plist if player_list_correct_counts[n]], key=lambda x: x[1], reverse=True)[:3])]
            h = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_correct_counts[n])) for n in plist if player_list_correct_counts[n]], key=lambda x: x[1])[:3])]
            z = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_vintages[n])) for n in plist if player_list_vintages[n]], key=lambda x: x[1], reverse=True)[:3])]
            b = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_vintages[n])) for n in plist if player_list_vintages[n]], key=lambda x: x[1])[:3])]
            
            pd.DataFrame([["Top 3 Easiest Lists"]] + e).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+1, startcol=0)
            pd.DataFrame([["Top 3 Hardest Lists"]] + h).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+1, startcol=1)
            pd.DataFrame([["Top 3 Zoomer Lists"]] + z).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+6, startcol=0)
            pd.DataFrame([["Top 3 Boomer Lists"]] + b).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+6, startcol=1)
            
            p_28 = sorted(plist, key=lambda x: player_two_eighths[x], reverse=True)[0]
            no_erig_pool = [n for n in plist if erigs_counts[n] == 0]
            best_no_erig = sorted(no_erig_pool, key=lambda x: (correct_counts[x]/song_participation[x]), reverse=True)[0] if no_erig_pool else "N/A"
            best_no_erig_gr = f"{correct_counts[best_no_erig]/song_participation[best_no_erig]:.2%}" if no_erig_pool else "N/A"

            pd.DataFrame([
                [f"Most 2/8s", f"{p_28} ({player_two_eighths[p_28]})"],
                [f"Highest GR with no erig", f"{best_no_erig} ({best_no_erig_gr})"]
            ]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+11, startcol=0)
            
            m_miss = max(player_missed_erigs, key=player_missed_erigs.get) if player_missed_erigs else "N/A"
            m_rev = max(player_reverse_erigs, key=player_reverse_erigs.get) if player_reverse_erigs else "N/A"
            pd.DataFrame([
                ["Top erig misser", f"{m_miss} ({player_missed_erigs.get(m_miss, 0)})"], 
                ["Top reverse erig collector", f"{m_rev} ({player_reverse_erigs.get(m_rev, 0)})"]
            ]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+14, startcol=0)

    # --- STYLING ---
    wb = load_workbook(out_path); ws_ps = wb["Player Stats"]; ws_extra = wb["Extra Stats"]
    bold, thin = Font(bold=True), Side(style='thin')
    outline = Border(left=thin, right=thin, top=thin, bottom=thin)
    green, red = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"), PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for row in ws_extra.iter_rows():
        for cell in row: cell.alignment = Alignment(horizontal='left')

    for row_offset in [12, 13]:
        ws_extra.cell(row=base_r+row_offset, column=1).alignment = Alignment(wrapText=True, horizontal='left', vertical='center')

    for col_name in pct_cols_p:
        col_idx = list(df_ps.columns).index(col_name) + 1
        vals = df_ps[col_name].dropna().unique()
        if len(vals) > 0:
            top3, bot3 = sorted(vals, reverse=True)[:3], sorted(vals)[:3]
            for r_idx, val in enumerate(df_ps[col_name], start=2):
                if pd.notnull(val):
                    if val in top3: ws_ps.cell(row=r_idx, column=col_idx).fill = green
                    elif val in bot3: ws_ps.cell(row=r_idx, column=col_idx).fill = red

    for r in range(1, 9): ws_extra.cell(row=r, column=1).font = bold
    ws_extra.cell(row=base_r+1, column=1).font = bold
    ws_extra.cell(row=base_r+1, column=1).border, ws_extra.cell(row=base_r+1, column=2).border = outline, outline

    bold_targets = ["Top 3 Easiest Lists", "Top 3 Hardest Lists", "Top 3 Zoomer Lists", "Top 3 Boomer Lists", "Most 2/8s", "Highest GR with no erig", "Top erig misser", "Top reverse erig collector", "Team with the most erigs", "Most zoomer team", "Most boomer team", "Team with the easiest lists", "Team with the hardest lists", "T1", "T2", "T3", "T4"]
    for row in ws_extra.iter_rows():
        for cell in row:
            if any(target in str(cell.value) for target in bold_targets): cell.font = bold
    
    for col in ws_ps.columns:
        max_l = max([len(str(cell.value)) for cell in col] + [0])
        ws_ps.column_dimensions[col[0].column_letter].width = max_l + 2
    
    ws_extra.column_dimensions['A'].width = 25
    for col in ['B', 'D', 'E', 'F']:
        max_l = max([len(str(cell.value)) for cell in ws_extra[col]] + [15])
        ws_extra.column_dimensions[col].width = max_l + 2

    wb.save(out_path)
    
    if messagebox.askyesno("Success", f"Exported to {out_name}.\nDo you want to delete all processed JSON files?"):
        for path in json_paths:
            try: os.remove(path)
            except: pass
        messagebox.showinfo("Cleanup", "JSON files deleted.")

if __name__ == "__main__":
    root = tk.Tk(); root.withdraw()
    process_files()
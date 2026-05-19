import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import numpy as np
import re
import os
from html import escape
from shutil import which
from datetime import datetime

try:
    from html2image import Html2Image
    from PIL import Image
except ImportError:
    Html2Image = None
    Image = None

# --- Configuration ---
EXCLUDED_TAGS = {
    "Female Protagonist", "Male Protagonist", "Primarily Female Cast", 
    "Primarily Male Cast", "School", "Heterosexual", "Primarily Teen Cast",
    "Ensemble Cast"
}

def extract_year(vintage_str):
    if not vintage_str: return None
    years = re.findall(r'\d{4}', str(vintage_str))
    if not years: return None
    year_val = float(years[0])
    season_map = {"winter": 0.00, "spring": 0.25, "summer": 0.50, "fall": 0.75}
    v_lower = str(vintage_str).lower()
    decimal = 0.0
    for season, val in season_map.items():
        if season in v_lower:
            decimal = val
            break
    return year_val + decimal

def get_browser():
    browser_paths = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        which("chrome"),
        which("msedge"),
    ]
    return next((path for path in browser_paths if path and os.path.exists(path)), None)

def trim_bottom_white(path_in):
    if Image is None:
        return
    img = Image.open(path_in).convert("RGBA")
    arr = np.array(img)
    rgb = arr[:, :, :3]
    alpha = arr[:, :, 3]
    non_white = np.any(rgb < 250, axis=2) & (alpha > 0)
    rows = np.where(non_white.any(axis=1))[0]
    if len(rows):
        img.crop((0, 0, img.width, rows[-1] + 8)).save(path_in)

def pct_text(value):
    return "N/A" if value is None or pd.isna(value) else f"{value:.2%}"

def number_text(value, decimals=2):
    return "N/A" if value is None or pd.isna(value) else f"{value:.{decimals}f}"

def parse_stat_cell(value, is_percent=False):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        has_percent = text.endswith("%")
        text = text.rstrip("%").replace(",", "")
        try:
            value = float(text)
        except ValueError:
            return None
        if has_percent:
            value /= 100
    elif isinstance(value, (int, float)):
        value = float(value)
    else:
        return None

    if is_percent and value > 1:
        value /= 100
    return value

def average_tour_blocks(ws, col_idx, is_percent=False):
    tour_values = []
    tour_averages = []
    for row_idx in range(2, ws.max_row + 1):
        value = parse_stat_cell(ws.cell(row_idx, col_idx).value, is_percent)
        if value is None:
            if tour_values:
                tour_averages.append(float(np.mean(tour_values)))
                tour_values = []
        else:
            tour_values.append(value)
    if tour_values:
        tour_averages.append(float(np.mean(tour_values)))
    return float(np.mean(tour_averages)) if tour_averages else None

def load_server_average_stats(script_dir):
    public_sheet_path = os.path.join(script_dir, "dependencies", "arpia", "public.xlsx")
    if not os.path.exists(public_sheet_path):
        return {}

    try:
        wb = load_workbook(public_sheet_path, data_only=True, read_only=False)
    except Exception:
        return {}

    sheet_map = {
        "watched": "Watched FL",
        "random": "Usual",
    }
    stats = {}
    for mode, sheet_name in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        stats[mode] = {
            "guess_rate": average_tour_blocks(ws, 4, is_percent=True),
            "attacker": average_tour_blocks(ws, 14),
            "blocker": average_tour_blocks(ws, 15),
        }
    return stats

def normalize_player_name(name):
    return str(name).strip().casefold()

def load_player_aliases(script_dir):
    preferred_path = os.path.join(script_dir, "dependencies", "arpia", "stats.xlsx")
    candidates = [preferred_path] if os.path.exists(preferred_path) else []
    if not candidates:
        arpia_dep_dir = os.path.join(script_dir, "dependencies", "arpia")
        if os.path.exists(arpia_dep_dir) and os.path.isdir(arpia_dep_dir):
            candidates = [
                os.path.join(arpia_dep_dir, f)
                for f in os.listdir(arpia_dep_dir)
                if f.lower().endswith(".xlsx") and not f.lower().startswith("export_")
            ]

    for path in candidates:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            if "IDs" not in wb.sheetnames:
                continue
            ws = wb["IDs"]
            alias_to_id = {}
            id_to_aliases = defaultdict(set)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) < 2:
                    continue
                player_name, player_id = row[0], row[1]
                if player_name is None or player_id is None:
                    continue
                norm_name = normalize_player_name(player_name)
                alias_to_id[norm_name] = player_id
                id_to_aliases[player_id].add(norm_name)
            if alias_to_id:
                return alias_to_id, id_to_aliases
        except Exception:
            continue
    return {}, defaultdict(set)

def resolve_player_name(input_name, available_pool, alias_to_id, id_to_aliases):
    exact_match = next((n for n in available_pool if normalize_player_name(n) == normalize_player_name(input_name)), None)
    if exact_match:
        return exact_match

    player_id = alias_to_id.get(normalize_player_name(input_name))
    if player_id is None:
        return None

    aliases = id_to_aliases.get(player_id, set())
    return next((n for n in available_pool if normalize_player_name(n) in aliases), None)

def medal_html(index):
    return ["&#x1F947;", "&#x1F948;", "&#x1F949;"][index]

def ranked_list_html(title, rows, formatter):
    lines = []
    for i, (name, value) in enumerate(rows[:3]):
        lines.append(f"<div>{medal_html(i)} {escape(str(name))} ({formatter(value)})</div>")
    while len(lines) < 3:
        lines.append("<div>&nbsp;</div>")
    return f"""
        <div class="podium">
            <div class="section-title">{escape(title)}</div>
            {''.join(lines)}
        </div>
    """

def line_plot_html(title, value, min_value, max_value, left_label, right_label, value_label, points=None, extra_marker=None):
    if value is None or max_value == min_value:
        pct = 0
    else:
        pct = max(0, min(100, ((value - min_value) / (max_value - min_value)) * 100))
    marker_html = ""
    stat_items = [f"<span><b>Tour Average</b><br>{value_label}</span>"]
    if extra_marker and extra_marker.get("value") is not None and max_value != min_value:
        marker_pct = max(0, min(100, ((extra_marker["value"] - min_value) / (max_value - min_value)) * 100))
        marker_html = f"""
            <div class="scale-marker server" style="left:{marker_pct}%"></div>
        """
        stat_items.append(f'<span class="server-stat"><b>Server Average</b><br>{pct_text(extra_marker["value"])}</span>')
    dots = []
    for name, point_value in points or []:
        if point_value is None or max_value == min_value:
            continue
        point_pct = max(1.5, min(98.5, ((point_value - min_value) / (max_value - min_value)) * 100))
        dots.append(f'<div class="plot-dot" style="left:{point_pct}%" title="{escape(str(name))}"></div>')
    return f"""
        <div class="scale-block">
            <div class="metric-line"><b>{escape(title)}</b></div>
            <div class="scale line-scale">
                {''.join(dots)}
                <div class="scale-marker tour" style="left:{pct}%"></div>
                {marker_html}
            </div>
            <div class="scale-labels">
                <span>{escape(left_label)}</span>
                <span>{escape(right_label)}</span>
            </div>
            <div class="avg-stats">{"".join(stat_items)}</div>
        </div>
    """

def hero_chart_html(title, rows, average_value, server_average=None):
    rows = rows or []
    max_value = max([value for _, _, value in rows] + [average_value or 0, server_average or 0, 1])
    avg_pct = max(0, min(100, ((average_value or 0) / max_value) * 100))
    server_pct = max(0, min(100, ((server_average or 0) / max_value) * 100))
    server_guide = ""
    server_axis = ""
    if server_average is not None:
        server_guide = f'<div class="guide server-guide" style="left:{server_pct}%"></div>'
        server_axis = f'<span class="server-axis"><b>Server Average</b><br>{number_text(server_average)}</span>'
    html_rows = []
    for tier, name, value in rows:
        width = max(2, min(100, (value / max_value) * 100)) if max_value else 0
        html_rows.append(f"""
            <div class="hero-row">
                <div class="tier">{escape(str(tier))}</div>
                <div class="hero-name">{escape(str(name))}</div>
                <div class="hero-bar-wrap"><div class="hero-bar" style="width:{width}%"></div></div>
                <div class="hero-value">{value}</div>
            </div>
        """)
    if not html_rows:
        html_rows.append('<div class="empty-note">No team data</div>')
    return f"""
        <div class="hero-chart">
            <div class="chart-title">{escape(title)}</div>
            <div class="hero-plot">
                <div class="hero-guides">
                    <div class="guide tour-guide" style="left:{avg_pct}%"></div>
                    {server_guide}
                </div>
                {''.join(html_rows)}
            </div>
            <div class="chart-axis">
                <span><b>Tour Average</b><br>{number_text(average_value)}</span>
                {server_axis}
            </div>
        </div>
    """

def save_extra_stats_image(data, output_dir, filename):
    if Html2Image is None:
        raise RuntimeError("html2image is not installed. Install it to export the Extra Stats PNG.")

    chanting_html = ""
    if data["has_chanting"]:
        chanting_html = f"""
            <div class="chanting">
                <div class="boxed-title">CHANTING STATS</div>
                <div class="two-col-row"><span>Total chanting songs played</span><span>{data['chanting_total']}</span></div>
                <div class="two-col-row"><span>Average chanting guess rate</span><span>{pct_text(data['chanting_gr'])}</span></div>
                <div class="chant-lists">
                    {ranked_list_html("Top 3 Chanting Lovers", data["chanting_lovers"], pct_text)}
                    {ranked_list_html("Top 3 Chanting Haters", data["chanting_haters"], pct_text)}
                </div>
            </div>
        """

    server_marker = None
    if data["server_average_gr"] is not None:
        server_marker = {
            "value": data["server_average_gr"],
            "label": f"Server Average\n{pct_text(data['server_average_gr'])}",
        }

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
    * {{ box-sizing: border-box; }}
    body {{
        margin: 0;
        background: white;
        color: black;
        font-family: Helvetica, Arial, sans-serif;
        font-size: 18px;
    }}
    .dashboard {{
        width: 1220px;
        min-height: 980px;
        padding: 0;
        display: grid;
        grid-template-columns: 530px 620px;
        gap: 36px;
    }}
    .left, .right {{ position: relative; }}
    .metric-line {{
        display: flex;
        justify-content: space-between;
        align-items: baseline;
        padding: 8px 18px 0 4px;
    }}
    .scale-block {{ margin-bottom: 18px; position: relative; }}
    .scale {{
        height: 44px;
        margin: 0 16px 0 6px;
        position: relative;
        overflow: visible;
    }}
    .line-scale::before {{
        content: "";
        position: absolute;
        left: 0;
        right: 0;
        top: 50%;
        border-top: 5px solid black;
        transform: translateY(-50%);
    }}
    .plot-dot {{
        position: absolute;
        top: 50%;
        width: 3px;
        height: 20px;
        transform: translate(-50%, -50%);
        border-radius: 0;
        background: #ed1c24;
        border: 0;
        box-shadow: none;
    }}
    .scale-marker {{
        position: absolute;
        top: 3px;
        bottom: 3px;
        width: 2px;
        background: black;
    }}
    .scale-marker.server {{
        width: 2px;
        background: #2563eb;
    }}
    .marker-label {{
        position: absolute;
        top: 47px;
        transform: translateX(-50%);
        text-align: center;
        white-space: pre;
        font-size: 16px;
        font-weight: bold;
    }}
    .server-label {{ color: #1d4ed8; }}
    .scale-labels {{
        position: relative;
        height: 22px;
        display: flex;
        justify-content: space-between;
        padding: 0 0 0 8px;
        font-size: 16px;
        font-weight: bold;
        margin-top: -2px;
    }}
    .avg-stats {{
        display: flex;
        gap: 28px;
        align-items: flex-start;
        margin: 0 0 10px 8px;
        font-size: 16px;
        font-weight: normal;
        line-height: 1.2;
    }}
    .avg-stats span {{
        text-align: left;
    }}
    .avg-stats b {{
        font-weight: bold;
    }}
    .server-stat {{ color: #1d4ed8; }}
    .boxed-title {{
        border: 2px solid black;
        height: 31px;
        line-height: 28px;
        padding-left: 5px;
        font-size: 23px;
        font-weight: bold;
        margin: 8px 0 10px;
    }}
    .podium-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        border-left: 1px solid #ccc;
        border-top: 1px solid #ccc;
    }}
    .podium {{
        min-height: 116px;
        border-right: 1px solid #ccc;
        border-bottom: 1px solid #ccc;
        padding: 0 4px 5px;
    }}
    .section-title {{
        font-size: 22px;
        font-weight: bold;
        margin-bottom: 4px;
    }}
    .podium div:not(.section-title) {{
        line-height: 25px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        font-family: Helvetica, Arial, sans-serif;
    }}
    .summary-table {{
        border-left: 1px solid #ccc;
        border-top: 1px solid #ccc;
        margin-top: 12px;
        font-size: 20px;
    }}
    .summary-table .two-col-row {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        border-bottom: 1px solid #ccc;
    }}
    .summary-table span {{
        padding: 4px;
        border-right: 1px solid #ccc;
    }}
    .summary-table span:first-child {{ font-weight: bold; }}
    .hero-chart {{
        height: 202px;
        border-top: 1px solid #ccc;
        margin-bottom: 10px;
        padding-top: 0;
    }}
    .chart-title {{
        height: 27px;
        border: 2px solid black;
        padding-left: 14px;
        font-weight: bold;
        line-height: 24px;
    }}
    .hero-plot {{
        margin-top: 0;
        position: relative;
        padding-top: 0;
    }}
    .hero-guides {{
        position: absolute;
        left: 147px;
        right: 45px;
        top: 0;
        bottom: 0;
        pointer-events: none;
        z-index: 3;
    }}
    .hero-row {{
        display: grid;
        grid-template-columns: 31px 116px 1fr 45px;
        height: 30px;
        align-items: center;
        font-weight: bold;
    }}
    .tier, .hero-name {{
        height: 30px;
        line-height: 29px;
        border-left: 1px solid black;
        border-bottom: 1px solid black;
        padding-left: 5px;
    }}
    .hero-name {{ border-right: 1px solid black; }}
    .hero-bar-wrap {{
        height: 30px;
        position: relative;
        border-bottom: 1px solid black;
        border-right: 1px solid black;
    }}
    .hero-bar {{
        height: 29px;
        background: #ed1c24;
        border-right: 2px solid black;
    }}
    .hero-value {{
        text-align: right;
        padding-right: 8px;
        font-family: Helvetica, Arial, sans-serif;
    }}
    .guide {{
        position: absolute;
        top: 0;
        bottom: 0;
        width: 2px;
        background: black;
        pointer-events: none;
    }}
    .server-guide {{
        background: #2563eb;
        width: 2px;
    }}
    .chart-axis {{
        display: flex;
        justify-content: flex-start;
        gap: 28px;
        height: 32px;
        margin-left: 0;
        margin-right: 45px;
        padding-left: 0;
        font-size: 16px;
        font-weight: bold;
    }}
    .chart-axis span {{
        text-align: center;
    }}
    .server-axis {{ color: #1d4ed8; }}
    .chanting {{
        margin-top: 18px;
        width: 552px;
    }}
    .chanting .boxed-title {{ margin-bottom: 0; }}
    .two-col-row {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        min-height: 30px;
        border-left: 1px solid #ccc;
        border-bottom: 1px solid #ccc;
    }}
    .two-col-row span {{
        padding: 4px 6px;
        border-right: 1px solid #ccc;
    }}
    .chant-lists {{
        margin-top: 28px;
        display: grid;
        grid-template-columns: 1fr 1fr;
        border-left: 1px solid #ccc;
        border-top: 1px solid #ccc;
    }}
    .empty-note {{ padding: 12px; color: #555; }}
</style>
</head>
<body>
        <div class="dashboard">
        <div class="left">
            {line_plot_html("Tour Average GR", data["tour_average_gr"], 0, 1, "0%", "100%", pct_text(data["tour_average_gr"]), data["gr_points"], server_marker)}
            <div class="boxed-title">WATCHED STATS</div>
            {line_plot_html("", data["watched_average"], 0, 8, "0.0", "8.0", number_text(data["watched_average"]), data["difficulty_points"])}
            <div class="podium-grid">
                {ranked_list_html("Top 3 Easiest Lists", data["easiest_lists"], number_text)}
                {ranked_list_html("Top 3 Hardest Lists", data["hardest_lists"], number_text)}
            </div>
            {line_plot_html("", data["vintage_average"], data["vintage_min"], data["vintage_max"], number_text(data["vintage_min"]), number_text(data["vintage_max"]), number_text(data["vintage_average"]), data["vintage_points"])}
            <div class="podium-grid">
                {ranked_list_html("Top 3 Zoomer Lists", data["zoomer_lists"], number_text)}
                {ranked_list_html("Top 3 Boomer Lists", data["boomer_lists"], number_text)}
            </div>
            <div class="summary-table">
                <div class="two-col-row"><span>Most 2/8s</span><span>{escape(data["most_two_eighths"])}</span></div>
                <div class="two-col-row"><span>Highest GR with no erig</span><span>{escape(data["best_no_erig"])}</span></div>
                <div class="two-col-row"><span>&nbsp;</span><span>&nbsp;</span></div>
                <div class="two-col-row"><span>Top erig misser</span><span>{escape(data["top_erig_misser"])}</span></div>
                <div class="two-col-row"><span>Top reverse erig collector</span><span>{escape(data["top_reverse_erig"])}</span></div>
            </div>
        </div>
        <div class="right">
            {hero_chart_html("Top Attacker", data["top_attackers"], data["attacker_average"], data["server_attacker_average"])}
            {hero_chart_html("Top Blocker", data["top_blockers"], data["blocker_average"], data["server_blocker_average"])}
            {chanting_html}
        </div>
    </div>
</body>
</html>
"""

    hti = Html2Image(
        size=(1228, 1120),
        browser_executable=get_browser(),
        custom_flags=[
            "--headless=new",
            "--hide-scrollbars",
            "--disable-gpu",
            "--force-device-scale-factor=1",
        ],
        output_path=output_dir
    )
    hti.screenshot(html_str=html, save_as=filename)
    save_path = os.path.join(output_dir, filename)
    trim_bottom_white(save_path)
    return save_path

# --- UI COMPONENTS ---

class TourSelectionDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Tour Selection")
        self.result = None

        tk.Label(self, text = "Select a tour to process the data for:", font = ("Arial", 10), padx = 20, pady = 15).pack()
        
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady = 10, padx = 20)
        
        for tour in ["Tour 0", "Tour 1", "Tour 2"]:
            ttk.Button(btn_frame, text = tour, command = lambda t = tour[-1]: self.set_result(t)).pack(side = tk.LEFT, padx = 5)
        
        self.protocol("WM_DELETE_WINDOW", self.on_cancel)
        self.grab_set()
        self.wait_window()

    def set_result(self, tour_id):
        self.result = tour_id
        self.destroy()

    def on_cancel(self):
        self.result = None
        self.destroy()

class SubSelectionDialog(tk.Toplevel):
    def __init__(self, parent, missing_roster):
        super().__init__(parent)
        self.title("Substitute Resolution")
        self.result = None
        tk.Label(self, text="Multiple roster members are missing.\nWhich player is being replaced by the substitute?", 
                 font=("Arial", 10), padx=20, pady=10).pack()
        self.listbox = tk.Listbox(self, height=len(missing_roster))
        self.listbox.pack(padx=20, pady=5, fill=tk.X)
        for m in missing_roster: self.listbox.insert(tk.END, m)
        ttk.Button(self, text="Confirm", command=self.on_confirm).pack(pady=10)
        self.grab_set(); self.wait_window()
        
    def on_confirm(self):
        sel = self.listbox.curselection()
        if sel: self.result = self.listbox.get(sel[0]); self.destroy()

class ManualMatchDialog(tk.Toplevel):
    def __init__(self, parent, unknown_name, available_pool):
        super().__init__(parent)
        self.title("Manual Match Required")
        self.result = None
        ttk.Label(self, text=f"Could not find match for: '{unknown_name}'", font=("Arial", 10, "bold")).pack(pady=10)
        self.listbox = tk.Listbox(self, height=15); self.listbox.pack(padx=10, fill=tk.BOTH)
        for name in sorted(available_pool): self.listbox.insert(tk.END, name)
        ttk.Button(self, text="Match Selected", command=self.on_match).pack(pady=10)
        self.grab_set(); self.wait_window()

    def on_match(self):
        sel = self.listbox.curselection()
        if sel: self.result = self.listbox.get(sel[0]); self.destroy()

class TourModeDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Tour Type")
        self.result = None
        self.resizable(False, False)
        ttk.Label(
            self,
            text="Which server averages should this export use?",
            font=("Arial", 10, "bold"),
            padding=(18, 14, 18, 8)
        ).pack()
        btn_frame = ttk.Frame(self, padding=(18, 4, 18, 16))
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="Watched", command=lambda: self.select("watched")).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="Usual / Random", command=lambda: self.select("random")).pack(side=tk.LEFT)
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        self.grab_set()
        self.wait_window()

    def select(self, mode):
        self.result = mode
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()

# --- CORE LOGIC ---

def process_files(tour_id):
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    json_dir = os.path.join(script_dir, "tours", tour_id, "jsons")
    codes_path = os.path.join(script_dir, "tours", tour_id, "codes.txt")
    out_dir = os.path.join(script_dir, "tours", tour_id, "output")
    out_path = os.path.join(out_dir, "Final.xlsx")

    if not os.path.exists(out_dir): 
        os.makedirs(out_dir)

    mode_dialog = TourModeDialog(None)
    server_average_mode = mode_dialog.result
    if server_average_mode is None:
        return
    
    # Load Chanting IDs
    chanting_ids = set()
    chanting_path = os.path.join(script_dir, "dependencies", "chanting", "chanting.txt")
    if os.path.exists(chanting_path):
        with open(chanting_path, "r") as f:
            for line in f:
                line = line.strip()
                if line: chanting_ids.add(line)

    json_paths = []
    while True:
        if os.path.exists(json_dir) and os.path.isdir(json_dir):
            json_paths = [os.path.join(json_dir, f) for f in os.listdir(json_dir) if f.endswith(".json")]
        
        if json_paths:
            break
        else:
            retry = messagebox.askyesno("Missing Files", f"There is no jsons folder detected or there are no JSON files in the folder '{json_dir}'. Lock in and press yes to re-run the script")
            if not retry:
                return

    all_known_players = set()
    for path in json_paths:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f); songs = data.get("songs", [])
                for s in songs:
                    for p in s.get("correctGuessPlayers", []): all_known_players.add(p)
                    for ls in s.get("listStates", []): all_known_players.add(ls["name"])
        except: continue

    raw_assignments = {}
    team_rosters = defaultdict(set)
    t1_lookup = {}
    use_teams = False
    server_average_gr = None
    alias_to_id, id_to_aliases = load_player_aliases(script_dir)

    codes_valid = False
    if os.path.exists(codes_path):
        with open(codes_path, "r", encoding="utf-8") as f:
            if f.read().strip():
                codes_valid = True

    if not codes_valid:
        if not messagebox.askyesno("Codes Missing", f"codes.txt is missing or empty in tours/{tour_id}/, skip team assignment phase?"):
            return
    else:
        with open(codes_path, "r", encoding="utf-8") as f:
            content = f.read()
        all_teams_data = []
        for line in content.strip().split('\n'):
            if line.lower().startswith(("average", "avg")):
                avg_match = re.search(r"(-?\d+(?:\.\d+)?)", line)
                if avg_match:
                    server_average_gr = float(avg_match.group(1))
                    if server_average_gr > 1:
                        server_average_gr /= 100
            matches = re.findall(r'([^\s(]+)\s*\([\d.]+\)', line)
            if matches: all_teams_data.append(matches[:4])

        if all_teams_data:
            use_teams = True
            available = list(all_known_players)
            for t_idx, members in enumerate(all_teams_data, 1):
                for i, p_in in enumerate(members):
                    tier = f"T{i+1}"
                    match = resolve_player_name(p_in, available, alias_to_id, id_to_aliases)
                    if not match:
                        d = ManualMatchDialog(None, p_in, available)
                        match = d.result
                    if match:
                        raw_assignments[match] = (t_idx, tier)
                        team_rosters[t_idx].add(match)
                        if match in available: available.remove(match)
                        if tier == "T1": t1_lookup[t_idx] = match

    correct_counts, song_participation = defaultdict(int), defaultdict(int)
    erigs_counts, player_reverse_erigs = defaultdict(int), defaultdict(int)
    player_two_eighths, player_points, player_blocks = defaultdict(int), defaultdict(int), defaultdict(int)
    player_type_correct, player_type_seen = defaultdict(lambda: defaultdict(int)), defaultdict(lambda: defaultdict(int))
    player_rigs, player_rigs_hit = defaultdict(int), defaultdict(int)
    all_song_vintages, all_song_difficulties = [], []
    total_correct_answers_sum, total_erigs = 0, 0
    genre_counter, tag_counter = Counter(), Counter()
    player_list_vintages, player_list_correct_counts = defaultdict(list), defaultdict(list) 
    player_missed_erigs, watched_only_valid = defaultdict(int), False
    team_correct_per_song = defaultdict(list)
    team_onlist_synergy, team_offlist_synergy, team_shared_rig_pct = defaultdict(list), defaultdict(list), defaultdict(list)
    
    total_songs_played = 0
    total_chanting_songs = 0
    player_chanting_correct = defaultdict(int)
    player_chanting_seen = defaultdict(int)
    chanting_correct_sum = 0

    for path in json_paths:
        with open(path, encoding="utf-8") as f: data = json.load(f); songs = data.get("songs", [])
        if not songs: continue
        
        raw_file_players = set()
        for song in songs:
            for p in song.get("correctGuessPlayers", []): raw_file_players.add(p)
            for ls in song.get("listStates", []): raw_file_players.add(ls["name"])
        
        final_file_members = set(raw_file_players)
        if use_teams:
            teams_in_file = set(raw_assignments[p][0] for p in raw_file_players if p in raw_assignments)
            for t_id in teams_in_file:
                roster = team_rosters[t_id]
                missing = [p for p in roster if p not in raw_file_players]
                if len([p for p in roster if p in raw_file_players]) == 3 and missing:
                    if len(missing) == 1: final_file_members.add(missing[0])
                    else:
                        d = SubSelectionDialog(None, missing)
                        if d.result: final_file_members.add(d.result)

        apply_rev = (len(final_file_members) % 2 == 0)
        max_songs = max(s.get("songNumber", 0) for s in songs)
        type_totals_this_file = defaultdict(int)

        for song in songs:
            total_songs_played += 1
            si = song.get("songInfo", {}); st = si.get("type")
            
            # Using annSongId for chanting matching
            ann_song_id = str(si.get("annSongId"))
            
            is_chanting = ann_song_id in chanting_ids
            if is_chanting: total_chanting_songs += 1

            if st in [1, 2, 3]: type_totals_this_file[st] += 1
            if isinstance(si.get("animeGenre"), list): genre_counter.update(si.get("animeGenre"))
            if isinstance(si.get("animeTags"), list):
                tag_counter.update([t for t in si.get("animeTags") if t not in EXCLUDED_TAGS])

            correct = set(song.get("correctGuessPlayers", []))
            ls = song.get("listStates", []); total_correct_answers_sum += len(correct)
            if is_chanting: chanting_correct_sum += len(correct)

            year, diff = extract_year(si.get("vintage")), si.get("animeDifficulty")
            if isinstance(diff, (int, float)): all_song_difficulties.append(diff)
            if year is not None: all_song_vintages.append(year)
            
            song_riggers = {p["name"] for p in ls}
            
            if use_teams:
                teams_in_this_file = list(set(raw_assignments[p][0] for p in raw_file_players if p in raw_assignments))
                if len(teams_in_this_file) == 2:
                    tA, tB = teams_in_this_file[0], teams_in_this_file[1]
                    for cur_t, opp_t in [(tA, tB), (tB, tA)]:
                        cur_correct = correct.intersection(team_rosters[cur_t])
                        opp_correct = correct.intersection(team_rosters[opp_t])
                        if not opp_correct:
                            for p in cur_correct: player_points[p] += 1
                        if len(cur_correct) == 1 and len(opp_correct) > 0:
                            player_blocks[list(cur_correct)[0]] += 1

                for t_id in teams_in_this_file:
                    roster = team_rosters[t_id]
                    correct_on_team = correct.intersection(roster)
                    team_correct_per_song[t_id].append(len(correct_on_team) / 4.0)
                    team_riggers = song_riggers.intersection(roster)
                    if team_riggers:
                        team_onlist_synergy[t_id].append(len(correct_on_team) / 4.0)
                        team_shared_rig_pct[t_id].append((len(team_riggers) - 1) / 3.0)
                    else: team_offlist_synergy[t_id].append(len(correct_on_team) / 4.0)

            if len(correct) == 2:
                for p in correct: player_two_eighths[p] += 1
            elif len(correct) == 1: 
                total_erigs += 1; erigs_counts[list(correct)[0]] += 1
            if apply_rev and len(final_file_members - correct) == 1:
                player_reverse_erigs[list(final_file_members - correct)[0]] += 1

            for name in final_file_members:
                if name in correct:
                    correct_counts[name] += 1
                    if st in [1, 2, 3]: player_type_correct[name][st] += 1
                    if is_chanting: player_chanting_correct[name] += 1
                if is_chanting: player_chanting_seen[name] += 1

            if ls:
                watched_only_valid = True
                for p in ls:
                    n = p["name"]; player_rigs[n] += 1
                    if n in correct: player_rigs_hit[n] += 1
                    if year is not None: player_list_vintages[n].append(year)
                    player_list_correct_counts[n].append(len(correct))
                    if len(correct) == 0: player_missed_erigs[n] += 1

        for name in final_file_members:
            song_participation[name] += max_songs
            for t in [1, 2, 3]: player_type_seen[name][t] += type_totals_this_file[t]

    p_rows = []
    for name in song_participation:
        total, correct = song_participation[name], correct_counts[name]
        t_id, tier = raw_assignments.get(name, ("Unassigned", "N/A"))
        t_name = t1_lookup.get(t_id, "Unknown")
        p_rows.append({
            "Team": t_name, "Tier": tier, "Player": name, 
            "Guess Rate": correct/total if total else 0, "Erigs 🔫": erigs_counts[name],
            "Points": player_points[name], "Blocks": player_blocks[name],
            "2/8s": player_two_eighths[name], "Rev. Erigs": player_reverse_erigs[name],
            "Song Count": total,
            "OP GR": player_type_correct[name][1]/player_type_seen[name][1] if player_type_seen[name][1] else np.nan,
            "ED GR": player_type_correct[name][2]/player_type_seen[name][2] if player_type_seen[name][2] else np.nan,
            "IN GR": player_type_correct[name][3]/player_type_seen[name][3] if player_type_seen[name][3] else np.nan,
            "Rigs": player_rigs[name], "Rigs Missed": player_rigs[name]-player_rigs_hit[name],
            "Onlist GR": player_rigs_hit[name]/player_rigs[name] if player_rigs[name] else np.nan,
            "Offlist GR": (correct-player_rigs_hit[name])/(total-player_rigs[name]) if (total-player_rigs[name]) else np.nan
        })
    df_ps = pd.DataFrame(p_rows).sort_values("Guess Rate", ascending=False)
    total_participation = sum(song_participation.values())
    avg_tour_gr = total_correct_answers_sum / total_participation if total_participation else 0

    df_tour = pd.DataFrame([
        ["Average Vintage", round(np.mean(all_song_vintages), 2) if all_song_vintages else "N/A"],
        ["Average Difficulty", round(np.mean(all_song_difficulties), 2) if all_song_difficulties else "N/A"],
        ["Average GR", f"{avg_tour_gr:.2%}"],
        ["Total Erigs", total_erigs],
        ["Total Rev. Erigs", sum(player_reverse_erigs.values())],
        ["Most Popular Genre", f"{genre_counter.most_common(1)[0][0]} ({genre_counter.most_common(1)[0][1]})" if genre_counter else "N/A"],
        ["Most Popular Tag", f"{tag_counter.most_common(1)[0][0]} ({tag_counter.most_common(1)[0][1]})" if tag_counter else "N/A"],
    ], columns=["TOUR STATS", ""])

    team_stat_rows, team_meta = [], []
    if use_teams:
        for t_id in sorted(team_correct_per_song.keys()):
            t_name = t1_lookup.get(t_id, f"Team {t_id}"); roster = team_rosters[t_id]
            t_v = [v for p in roster for v in player_list_vintages[p]]
            t_d = [v for p in roster for v in player_list_correct_counts[p]]
            team_stat_rows.append({"TEAM STATS": t_name, "Avg. Correct": np.mean(team_correct_per_song[t_id]), "Onlist Synergy": np.mean(team_onlist_synergy[t_id]) if team_onlist_synergy[t_id] else 0, "Offlist Synergy": np.mean(team_offlist_synergy[t_id]) if team_offlist_synergy[t_id] else 0, "Shared Rigs": np.mean(team_shared_rig_pct[t_id]) if team_shared_rig_pct[t_id] else 0})
            team_meta.append({"name": t_name, "erigs": sum(erigs_counts[p] for p in roster), "vintage": np.mean(t_v) if t_v else 0, "diff": np.mean(t_d) if t_d else 0})
    
    df_team_stats = pd.DataFrame(team_stat_rows)
    if not df_team_stats.empty:
        df_team_stats = df_team_stats.sort_values("Avg. Correct", ascending=False)

    tier_hero_rows = []
    tier_attackers, tier_blockers = {}, {}
    if use_teams:
        for tier in ["T1", "T2", "T3", "T4"]:
            tp = [p for p, attr in raw_assignments.items() if attr[1] == tier]
            if tp:
                bp = max(tp, key=lambda x: player_points[x]); bb = max(tp, key=lambda x: player_blocks[x])
                tier_attackers[tier] = (bp, player_points[bp])
                tier_blockers[tier] = (bb, player_blocks[bb])
                tier_hero_rows.append([tier, f"{bp} ({player_points[bp]})", f"{bb} ({player_blocks[bb]})"])
    df_tier_heroes = pd.DataFrame(tier_hero_rows, columns=["Tier", "Top Attacker", "Top Blocker"])

    plist = list(song_participation.keys())
    diff_data = [(n, np.mean(player_list_correct_counts[n])) for n in plist if player_list_correct_counts[n]]
    vint_data = [(n, np.mean(player_list_vintages[n])) for n in plist if player_list_vintages[n]]
    all_list_correct_counts = [v for values in player_list_correct_counts.values() for v in values]
    all_list_vintages = [v for values in player_list_vintages.values() for v in values]

    no_erig_pool = [n for n in plist if erigs_counts[n] == 0]
    best_no_erig = sorted(no_erig_pool, key=lambda x: (correct_counts[x]/song_participation[x]), reverse=True)[0] if no_erig_pool else "N/A"
    best_no_erig_gr = f"{correct_counts[best_no_erig]/song_participation[best_no_erig]:.2%}" if no_erig_pool else "N/A"
    p_28 = sorted(plist, key=lambda x: player_two_eighths[x], reverse=True)[0] if plist else "N/A"
    m_miss = max(player_missed_erigs, key=player_missed_erigs.get) if player_missed_erigs else "N/A"
    m_rev = max(player_reverse_erigs, key=player_reverse_erigs.get) if player_reverse_erigs else "N/A"

    chan_pct = total_chanting_songs / total_songs_played if total_songs_played else 0
    avg_chan_gr = (chanting_correct_sum / (total_chanting_songs * len(song_participation))) if (total_chanting_songs and song_participation) else 0
    chan_plist = [n for n in song_participation.keys() if player_chanting_seen[n] > 0]
    chan_rates = [(n, player_chanting_correct[n]/player_chanting_seen[n]) for n in chan_plist]
    server_average_stats = load_server_average_stats(script_dir)
    selected_server_stats = server_average_stats.get(server_average_mode, {})
    image_server_gr = selected_server_stats.get("guess_rate", server_average_gr)
    image_server_attacker = selected_server_stats.get("attacker")
    image_server_blocker = selected_server_stats.get("blocker")
    vintage_min = min([v for _, v in vint_data]) if vint_data else 0
    vintage_max = max([v for _, v in vint_data]) if vint_data else 1
    if vintage_min == vintage_max:
        vintage_min -= 1
        vintage_max += 1

    extra_image_data = {
        "tour_average_gr": avg_tour_gr,
        "server_average_gr": image_server_gr,
        "gr_points": [(row["Player"], row["Guess Rate"]) for row in p_rows],
        "watched_average": np.mean(all_list_correct_counts) if all_list_correct_counts else None,
        "difficulty_points": diff_data,
        "easiest_lists": sorted(diff_data, key=lambda x: x[1], reverse=True)[:3],
        "hardest_lists": sorted(diff_data, key=lambda x: x[1])[:3],
        "vintage_average": np.mean(all_list_vintages) if all_list_vintages else None,
        "vintage_min": vintage_min,
        "vintage_max": vintage_max,
        "vintage_points": vint_data,
        "zoomer_lists": sorted(vint_data, key=lambda x: x[1], reverse=True)[:3],
        "boomer_lists": sorted(vint_data, key=lambda x: x[1])[:3],
        "most_two_eighths": f"{p_28} ({player_two_eighths[p_28]})" if p_28 != "N/A" else "N/A",
        "best_no_erig": f"{best_no_erig} ({best_no_erig_gr})" if no_erig_pool else "N/A",
        "top_erig_misser": f"{m_miss} ({player_missed_erigs.get(m_miss, 0)})" if m_miss != "N/A" else "N/A",
        "top_reverse_erig": f"{m_rev} ({player_reverse_erigs.get(m_rev, 0)})" if m_rev != "N/A" else "N/A",
        "top_attackers": [(tier, tier_attackers[tier][0], tier_attackers[tier][1]) for tier in ["T1", "T2", "T3", "T4"] if tier in tier_attackers],
        "top_blockers": [(tier, tier_blockers[tier][0], tier_blockers[tier][1]) for tier in ["T1", "T2", "T3", "T4"] if tier in tier_blockers],
        "attacker_average": np.mean([player_points[n] for n in plist]) if plist else 0,
        "blocker_average": np.mean([player_blocks[n] for n in plist]) if plist else 0,
        "server_attacker_average": image_server_attacker,
        "server_blocker_average": image_server_blocker,
        "has_chanting": bool(chanting_ids),
        "chanting_total": f"{total_chanting_songs} ({chan_pct:.2%})",
        "chanting_gr": avg_chan_gr,
        "chanting_lovers": sorted(chan_rates, key=lambda x: x[1], reverse=True)[:3],
        "chanting_haters": sorted(chan_rates, key=lambda x: x[1])[:3],
    }

    df_display = df_ps.copy()
    pct_cols_p = ["Guess Rate", "OP GR", "ED GR", "IN GR", "Onlist GR", "Offlist GR"]
    for c in pct_cols_p: df_display[c] = df_display[c].apply(lambda x: f"{x:.2%}" if pd.notnull(x) else "N/A")

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_display.to_excel(writer, sheet_name="Player Stats", index=False)
        df_tour.to_excel(writer, sheet_name="Extra Stats", index=False)
        
        erig_r = 0 
        if use_teams and not df_team_stats.empty:
            df_team_display = df_team_stats.copy()
            pct_cols_t = ["Avg. Correct", "Onlist Synergy", "Offlist Synergy", "Shared Rigs"]
            for c in pct_cols_t: df_team_display[c] = df_team_display[c].apply(lambda x: f"{x:.2%}")
            df_team_display.to_excel(writer, sheet_name="Extra Stats", index=False, startcol=3)
            
            erig_r = len(df_team_stats) + 2
            if team_meta:
                m_e = sorted(team_meta, key=lambda x: x['erigs'], reverse=True)[0]
                m_z = sorted(team_meta, key=lambda x: x['vintage'], reverse=True)[0]
                m_b = sorted(team_meta, key=lambda x: x['vintage'])[0]
                m_ea = sorted(team_meta, key=lambda x: x['diff'], reverse=True)[0]
                m_h = sorted(team_meta, key=lambda x: x['diff'])[0]
                df_team_super = pd.DataFrame([
                    ["Team with the most erigs", f"{m_e['name']} ({m_e['erigs']})"],
                    ["Most zoomer team", f"{m_z['name']} ({round(m_z['vintage'], 2)})"],
                    ["Most boomer team", f"{m_b['name']} ({round(m_b['vintage'], 2)})"],
                    ["Team with the easiest lists", f"{m_ea['name']} ({round(m_ea['diff'], 2)})"],
                    ["Team with the hardest lists", f"{m_h['name']} ({round(m_h['diff'], 2)})"],
                ])
                df_team_super.iloc[0:1].to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=erig_r, startcol=3)
                df_team_super.iloc[1:].to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=erig_r + 1, startcol=3)
            
            if not df_tier_heroes.empty:
                df_tier_heroes.to_excel(writer, sheet_name="Extra Stats", index=False, startrow=erig_r + 6, startcol=3)

        chan_base_r = erig_r + 12 if use_teams else len(df_tour) + 3
        if chanting_ids:
            pd.DataFrame([["CHANTING STATS"]]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=chan_base_r, startcol=3)
            chan_pct = total_chanting_songs / total_songs_played if total_songs_played else 0
            avg_chan_gr = (chanting_correct_sum / (total_chanting_songs * len(song_participation))) if (total_chanting_songs and song_participation) else 0
            
            df_chan_sum = pd.DataFrame([
                ["Total chanting songs played", f"{total_chanting_songs} ({chan_pct:.2%})"],
                ["Average chanting guess rate", f"{avg_chan_gr:.2%}"]
            ])
            df_chan_sum.to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=chan_base_r + 1, startcol=3)

            chan_plist = [n for n in song_participation.keys() if player_chanting_seen[n] > 0]
            chan_rates = [(n, player_chanting_correct[n]/player_chanting_seen[n]) for n in chan_plist]
            high_chan = [[f"{['🥇','🥈','🥉'][i]} {p} ({v:.2%})"] for i, (p, v) in enumerate(sorted(chan_rates, key=lambda x: x[1], reverse=True)[:3])]
            low_chan = [[f"{['🥇','🥈','🥉'][i]} {p} ({v:.2%})"] for i, (p, v) in enumerate(sorted(chan_rates, key=lambda x: x[1])[:3])]
            
            pd.DataFrame([["Top 3 Chanting Lovers"]] + high_chan).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=chan_base_r + 4, startcol=3)
            pd.DataFrame([["Top 3 Chanting Haters"]] + low_chan).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=chan_base_r + 4, startcol=4)

        base_r = len(df_tour) + 2
        pd.DataFrame([["WATCHED STATS"]]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r)
        
        if watched_only_valid:
            plist = list(song_participation.keys())
            e = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_correct_counts[n])) for n in plist if player_list_correct_counts[n]], key=lambda x: x[1], reverse=True)[:3])]
            h = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_correct_counts[n])) for n in plist if player_list_correct_counts[n]], key=lambda x: x[1])[:3])]
            z = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_vintages[n])) for n in plist if player_list_vintages[n]], key=lambda x: x[1], reverse=True)[:3])]
            b = [[f"{['🥇','🥈','🥉'][i]} {p} ({round(v, 2)})"] for i, (p, v) in enumerate(sorted([(n, np.mean(player_list_vintages[n])) for n in plist if player_list_vintages[n]], key=lambda x: x[1])[:3])]
            
            pd.DataFrame([["Top 3 Easiest Lists"]] + e).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+1, startcol=0)
            pd.DataFrame([["Top 3 Hardest Lists"]] + h).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+1, startcol=1)
            pd.DataFrame([["Top 3 Zoomer Lists"]] + z).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+6, startcol=0)
            pd.DataFrame([["Top 3 Boomer Lists"]] + b).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+6, startcol=1)
            
            p_28 = sorted(plist, key=lambda x: player_two_eighths[x], reverse=True)[0]
            no_erig_pool = [n for n in plist if erigs_counts[n] == 0]
            best_no_erig = sorted(no_erig_pool, key=lambda x: (correct_counts[x]/song_participation[x]), reverse=True)[0] if no_erig_pool else "N/A"
            best_no_erig_gr = f"{correct_counts[best_no_erig]/song_participation[best_no_erig]:.2%}" if no_erig_pool else "N/A"

            pd.DataFrame([
                [f"Most 2/8s", f"{p_28} ({player_two_eighths[p_28]})"],
                [f"Highest GR with no erig", f"{best_no_erig} ({best_no_erig_gr})"]
            ]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+11, startcol=0)
            
            m_miss = max(player_missed_erigs, key=player_missed_erigs.get) if player_missed_erigs else "N/A"
            m_rev = max(player_reverse_erigs, key=player_reverse_erigs.get) if player_reverse_erigs else "N/A"
            pd.DataFrame([
                ["Top erig misser", f"{m_miss} ({player_missed_erigs.get(m_miss, 0)})"], 
                ["Top reverse erig collector", f"{m_rev} ({player_reverse_erigs.get(m_rev, 0)})"]
            ]).to_excel(writer, sheet_name="Extra Stats", index=False, header=False, startrow=base_r+14, startcol=0)

    wb = load_workbook(out_path); ws_ps = wb["Player Stats"]; ws_extra = wb["Extra Stats"]
    bold, thin = Font(bold=True), Side(style='thin')
    outline = Border(left=thin, right=thin, top=thin, bottom=thin)
    green, red = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"), PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    for row in ws_extra.iter_rows():
        for cell in row: cell.alignment = Alignment(horizontal='left')

    for col_name in pct_cols_p:
        col_idx = list(df_ps.columns).index(col_name) + 1
        vals = df_ps[col_name].dropna().unique()
        if len(vals) > 0:
            top3, bot3 = sorted(vals, reverse=True)[:3], sorted(vals)[:3]
            for r_idx, val in enumerate(df_ps[col_name], start=2):
                if pd.notnull(val):
                    if val in top3: ws_ps.cell(row=r_idx, column=col_idx).fill = green
                    elif val in bot3: ws_ps.cell(row=r_idx, column=col_idx).fill = red

    for r in range(1, 9): ws_extra.cell(row=r, column=1).font = bold
    ws_extra.cell(row=base_r+1, column=1).font = bold
    ws_extra.cell(row=base_r+1, column=1).border, ws_extra.cell(row=base_r+1, column=2).border = outline, outline

    if chanting_ids:
        ws_extra.cell(row=chan_base_r+1, column=4).font = bold
        ws_extra.cell(row=chan_base_r+1, column=4).border = outline
        ws_extra.cell(row=chan_base_r+1, column=5).border = outline
        
        ws_extra.cell(row=chan_base_r+5, column=4).font = bold
        ws_extra.cell(row=chan_base_r+5, column=5).font = bold

    bold_targets = ["Top 3 Easiest Lists", "Top 3 Hardest Lists", "Top 3 Zoomer Lists", "Top 3 Boomer Lists", "Most 2/8s", "Highest GR with no erig", "Top erig misser", "Top reverse erig collector", "Team with the most erigs", "Most zoomer team", "Most boomer team", "Team with the easiest lists", "Team with the hardest lists", "T1", "T2", "T3", "T4", "CHANTING STATS", "Top 3 High Chanting GR", "Top 3 Low Chanting GR"]
    for row in ws_extra.iter_rows():
        for cell in row:
            if any(target in str(cell.value) for target in bold_targets): cell.font = bold
    
    for col in ws_ps.columns:
        max_l = max([len(str(cell.value)) for cell in col] + [0])
        ws_ps.column_dimensions[col[0].column_letter].width = max_l + 2
    
    ws_extra.column_dimensions['A'].width = 25
    ws_extra.column_dimensions['C'].width = 4.2
    for col in ['B', 'D', 'E', 'F', 'G', 'H']:
        max_l = max([len(str(cell.value)) for cell in ws_extra[col]] + [15])
        ws_extra.column_dimensions[col].width = max_l + 2

    wb.save(out_path)
    image_name = "Final-4.png"
    image_status = f"\nExtra stats image exported to tours/{tour_id}/output/{image_name}."
    try:
        save_extra_stats_image(extra_image_data, out_dir, image_name)
    except Exception as exc:
        image_status = f"\nExtra stats image could not be exported: {exc}"
        messagebox.showwarning("Image Export Warning", image_status.strip())
    
    messagebox.showinfo("Success", f"Saved to tours/{tour_id}/output/Final.xlsx.{image_status}")

if __name__ == "__main__":
    root = tk.Tk(); root.withdraw()
    selector = TourSelectionDialog(root)
    if selector.result is not None:
        process_files(selector.result)
    else:
        print("Selection cancelled")